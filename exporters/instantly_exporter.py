"""
CSV/Excel export functionality for Instantly.ai cold email campaigns.
Exports enriched and personalized lead data in Instantly.ai's required format.
"""

import csv
import io
from datetime import datetime
from typing import List, Dict, Optional


def parse_name(business_name: str) -> tuple:
    """
    Parse business name into first_name, last_name for owner.
    Returns: (first_name, last_name)

    Examples:
    - "Smith's Nursery" -> ("Smith's", "Nursery")
    - "ABC Greenhouse Inc" -> ("ABC", "Greenhouse")
    - "Green Valley" -> ("Green", "Valley")
    """
    if not business_name:
        return ('', '')

    # Remove common suffixes
    suffixes = [' Inc', ' LLC', ' Corp', ' Co', ' Company', ' Ltd', ' LTD']
    cleaned = business_name
    for suffix in suffixes:
        if cleaned.endswith(suffix):
            cleaned = cleaned[:-len(suffix)].strip()

    # Split on first space
    parts = cleaned.split(' ', 1)
    if len(parts) == 2:
        return (parts[0], parts[1])
    elif len(parts) == 1:
        return (parts[0], '')
    else:
        return ('', '')


def format_location(address: Optional[str], city: Optional[str], state: Optional[str]) -> str:
    """
    Format location field from address components.
    Returns: "City, State" or "Address" or ""
    """
    if city and state:
        return f"{city}, {state}"
    elif address:
        return address
    else:
        return ''


def export_to_csv(leads: List[Dict], include_metadata: bool = True) -> str:
    """
    Export leads to CSV format for Instantly.ai.

    Args:
        leads: List of lead dictionaries from database
        include_metadata: Include tier, score, business_type columns (default True)

    Returns:
        CSV string ready for download

    CSV Columns (Instantly.ai format):
    - email: Owner/business email (required)
    - first_name: Parsed from business name
    - last_name: Parsed from business name
    - company_name: Business name
    - phone: Business phone number
    - website: Business website URL
    - location: "City, State" format
    - custom_line: AI-generated personalization (maps to {{Personalization}})

    Optional metadata columns (for reference):
    - tier: Lead tier (A/B/C/U)
    - score: Lead score (0-100)
    - business_type: Gemini-classified type
    """

    # Build CSV in memory
    output = io.StringIO()

    # Define columns
    if include_metadata:
        fieldnames = [
            'email', 'first_name', 'last_name', 'company_name',
            'phone', 'website', 'location', 'custom_line',
            'tier', 'score', 'business_type'
        ]
    else:
        fieldnames = [
            'email', 'first_name', 'last_name', 'company_name',
            'phone', 'website', 'location', 'custom_line'
        ]

    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()

    for lead in leads:
        # Parse name
        first_name, last_name = parse_name(lead.get('business_name', ''))

        # Build row
        row = {
            'email': lead.get('owner_email', ''),
            'first_name': first_name,
            'last_name': last_name,
            'company_name': lead.get('business_name', ''),
            'phone': lead.get('phone', ''),
            'website': lead.get('website', ''),
            'location': format_location(
                lead.get('address'),
                lead.get('city'),
                lead.get('state')
            ),
            'custom_line': lead.get('custom_line', '')
        }

        # Add metadata if requested
        if include_metadata:
            # Use tier_override if exists, otherwise tier
            tier = lead.get('tier_override') or lead.get('tier', '')
            row['tier'] = tier
            row['score'] = lead.get('score', '')
            row['business_type'] = lead.get('business_type', '')

        writer.writerow(row)

    return output.getvalue()


def export_to_excel(leads: List[Dict], include_metadata: bool = True):
    """
    Export leads to Excel format (XLSX) for backup/analysis.

    Args:
        leads: List of lead dictionaries from database
        include_metadata: Include tier, score, business_type columns (default True)

    Returns:
        BytesIO object containing Excel file

    Note: Requires openpyxl package. Install with: pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Export"

    # Define columns
    if include_metadata:
        headers = [
            'Email', 'First Name', 'Last Name', 'Company Name',
            'Phone', 'Website', 'Location', 'Custom Line',
            'Tier', 'Score', 'Business Type'
        ]
    else:
        headers = [
            'Email', 'First Name', 'Last Name', 'Company Name',
            'Phone', 'Website', 'Location', 'Custom Line'
        ]

    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_num, lead in enumerate(leads, 2):
        first_name, last_name = parse_name(lead.get('business_name', ''))
        tier = lead.get('tier_override') or lead.get('tier', '')

        row_data = [
            lead.get('owner_email', ''),
            first_name,
            last_name,
            lead.get('business_name', ''),
            lead.get('phone', ''),
            lead.get('website', ''),
            format_location(lead.get('address'), lead.get('city'), lead.get('state')),
            lead.get('custom_line', '')
        ]

        if include_metadata:
            row_data.extend([
                tier,
                lead.get('score', ''),
                lead.get('business_type', '')
            ])

        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1])
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def get_export_filename(format: str = 'csv', tier_filter: Optional[str] = None) -> str:
    """
    Generate filename for export.

    Args:
        format: 'csv' or 'xlsx'
        tier_filter: Optional tier filter (e.g., 'A', 'AB', 'ABC')

    Returns:
        Filename string (e.g., "leads_export_tierA_2024-01-21.csv")
    """
    timestamp = datetime.now().strftime('%Y-%m-%d')

    if tier_filter:
        tier_part = f"_tier{tier_filter}"
    else:
        tier_part = ""

    return f"leads_export{tier_part}_{timestamp}.{format}"
