#!/usr/bin/env python3
"""
Import Oregon OLCC (now OCM) cannabis producer/cultivator licenses into registries table.

Source: Oregon OLCC Tableau dashboard / direct CSV export
URL: https://www.oregon.gov/olcc/marijuana/Pages/Recreational-Marijuana-Licensee-Reports.aspx
Data: Tableau endpoint at data.olcc.state.or.us

Segment: cannabis_grower
Registry source: or_olcc
"""

import argparse
import json
import os
import sqlite3
import sys
import time
import re
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook
from io import BytesIO

# Ensure project root is in path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from database.models import get_db_connection, migrate_db

REGISTRY_SOURCE = 'or_olcc'
SEGMENT = 'cannabis_grower'

# Producer/cultivator license type keywords to filter
CULTIVATOR_TYPES = {
    'producer', 'cultivator', 'craft cannabis', 'craft marijuana',
    'producer tier i', 'producer tier ii', 'marijuana producer',
    'cannabis producer',
}

# OLCC Tableau data URL - try multiple approaches
OFFICIAL_XLSX_URL = (
    'https://www.oregon.gov/olcc/marijuana/Documents/Cannabis-Business-Licenses-All.xlsx'
)

TABLEAU_CSV_URL = (
    'https://data.olcc.state.or.us/t/OLCCPublic/views/'
    'CannabisBusinessLicensesEndorsements/CannabisLicensesEndorsements.csv'
)

# Alternative: OLCC CAMP eLicensing public API
ELICENSING_BASE = 'https://camp.olcc.online/prod/api/public'

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/120.0.0.0 Safari/537.36'
    ),
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}


def fetch_olcc_via_official_xlsx(xlsx_url: str = OFFICIAL_XLSX_URL):
    """Fetch OLCC public statewide business license workbook and filter cultivators."""
    print('[OR OLCC] Trying official Oregon.gov XLSX export...')
    try:
        r = requests.get(xlsx_url, timeout=90, headers=HEADERS)
        r.raise_for_status()
        wb = load_workbook(filename=BytesIO(r.content), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return []

        idx = {str(col).strip().lower(): i for i, col in enumerate(header) if col is not None}

        def g(row, *names):
            for n in names:
                i = idx.get(n)
                if i is not None and i < len(row) and row[i] is not None:
                    return str(row[i]).strip()
            return ''

        records = []
        for row in rows:
            license_type = g(row, 'license type', 'license_type').lower()
            if not any(t in license_type for t in CULTIVATOR_TYPES):
                continue

            physical = g(row, 'physicaladdress', 'physical address', 'address')
            city = ''
            zip_code = ''
            m = re.search(r'\b([A-Za-z .\'-]+)\s+OR\s+(\d{5}(?:-\d{4})?)', physical)
            if m:
                city = m.group(1).strip()
                zip_code = m.group(2).strip()

            records.append({
                'business_name': g(row, 'business name', 'trade name') or g(row, 'business licenses'),
                'license_number': g(row, 'license number'),
                'license_type': license_type,
                'license_status': 'active',
                'city': city,
                'zip': zip_code,
                'county': g(row, 'county'),
                'address': physical,
                'raw_data': json.dumps({str(header[i]): ('' if i >= len(row) or row[i] is None else str(row[i])) for i in range(len(header))}),
            })

        print(f'[OR OLCC] Official XLSX fetched {len(records)} cultivator records')
        return records
    except Exception as e:
        print(f'[OR OLCC] Official XLSX error: {e}')
        return []


def fetch_olcc_via_tableau():
    """Try to get data from OLCC Tableau export."""
    session = requests.Session()
    session.headers.update(HEADERS)

    print('[OR OLCC] Trying Tableau CSV endpoint...')
    try:
        # First get the view to establish session
        view_url = (
            'https://data.olcc.state.or.us/t/OLCCPublic/views/'
            'CannabisBusinessLicensesEndorsements/CannabisLicensesEndorsements'
        )
        r = session.get(view_url, timeout=30)

        # Now try CSV download
        csv_url = view_url + '.csv'
        r2 = session.get(csv_url, timeout=60, params={':showVizHome': 'no'})

        if r2.status_code == 200 and 'License Number' in r2.text[:2000]:
            print(f'[OR OLCC] Tableau CSV: got {len(r2.text)} bytes')
            return parse_tableau_csv(r2.text)
        else:
            print(f'[OR OLCC] Tableau CSV returned {r2.status_code}, content-type: {r2.headers.get("content-type")}')
            return []
    except Exception as e:
        print(f'[OR OLCC] Tableau error: {e}')
        return []


def parse_tableau_csv(csv_text):
    """Parse CSV from Tableau export."""
    import csv
    import io

    records = []
    reader = csv.DictReader(io.StringIO(csv_text))

    for row in reader:
        # Map field names (Tableau may use different names)
        license_type = (
            row.get('License Type', '')
            or row.get('license_type', '')
            or row.get('License_Type', '')
        ).lower()

        # Filter for producers/cultivators
        is_cultivator = any(t in license_type for t in CULTIVATOR_TYPES)
        if not is_cultivator:
            continue

        records.append({
            'business_name': (
                row.get('Trade Name', '')
                or row.get('Business Name', '')
                or row.get('trade_name', '')
                or ''
            ).strip(),
            'license_number': (
                row.get('License Number', '')
                or row.get('license_number', '')
                or ''
            ).strip(),
            'license_type': license_type,
            'license_status': (
                row.get('Status', '')
                or row.get('status', '')
                or 'active'
            ).strip(),
            'city': (
                row.get('City', '')
                or row.get('city', '')
                or ''
            ).strip(),
            'zip': (
                row.get('Zip', '')
                or row.get('zip', '')
                or row.get('Zip Code', '')
                or ''
            ).strip(),
            'county': (
                row.get('County', '')
                or row.get('county', '')
                or ''
            ).strip(),
            'address': (
                row.get('Address', '')
                or row.get('address', '')
                or ''
            ).strip(),
            'raw_data': json.dumps(dict(row)),
        })

    return records


def fetch_olcc_via_elicensing():
    """Try OLCC CAMP eLicensing public search API."""
    print('[OR OLCC] Trying CAMP eLicensing API...')
    session = requests.Session()
    session.headers.update(HEADERS)

    records = []
    page = 0
    page_size = 100

    # Producer license type IDs — these are Oregon-specific
    license_type_names = ['Marijuana Producer', 'Cannabis Producer']

    for license_type_name in license_type_names:
        page = 0
        while True:
            try:
                params = {
                    'licenseTypeName': license_type_name,
                    'status': 'Active',
                    'page': page,
                    'pageSize': page_size,
                }
                r = session.get(
                    f'{ELICENSING_BASE}/licenses',
                    params=params,
                    timeout=30,
                )

                if r.status_code == 429:
                    print('[OR OLCC] Rate limited, sleeping 60s...')
                    time.sleep(60)
                    continue
                elif r.status_code != 200:
                    print(f'[OR OLCC] eLicensing returned {r.status_code}')
                    break

                data = r.json()
                items = data if isinstance(data, list) else data.get('results', data.get('data', []))

                if not items:
                    break

                for item in items:
                    records.append({
                        'business_name': (
                            item.get('tradeName', '')
                            or item.get('businessName', '')
                            or item.get('name', '')
                            or ''
                        ).strip(),
                        'license_number': str(item.get('licenseNumber', '') or item.get('license', '') or '').strip(),
                        'license_type': license_type_name.lower(),
                        'license_status': str(item.get('status', 'active')).lower(),
                        'city': (item.get('city', '') or '').strip(),
                        'zip': str(item.get('zip', '') or item.get('zipCode', '') or '').strip(),
                        'county': (item.get('county', '') or '').strip(),
                        'address': (item.get('address', '') or '').strip(),
                        'raw_data': json.dumps(item),
                    })

                if len(items) < page_size:
                    break

                page += 1
                time.sleep(1)

            except Exception as e:
                print(f'[OR OLCC] eLicensing error page {page}: {e}')
                break

    print(f'[OR OLCC] eLicensing fetched {len(records)} records')
    return records


def fetch_olcc_via_html_scrape():
    """
    Scrape OLCC CAMP portal HTML search results.
    Fallback when APIs are unavailable.
    """
    print('[OR OLCC] Trying CAMP portal HTML scrape...')
    from bs4 import BeautifulSoup

    session = requests.Session()
    session.headers.update(HEADERS)

    records = []
    base_url = 'https://camp.olcc.online/prod/webui/'

    try:
        # Load the main search page
        r = session.get(base_url, timeout=30)
        if r.status_code != 200:
            print(f'[OR OLCC] CAMP portal returned {r.status_code}')
            return []

        # This is a JS-heavy portal — HTML scraping won't work well here
        # Return empty and let caller handle fallback
        print('[OR OLCC] CAMP portal requires JS rendering, skipping HTML scrape')
        return []

    except Exception as e:
        print(f'[OR OLCC] HTML scrape error: {e}')
        return []


def fetch_records(official_xlsx_url: str = ''):
    """
    Main fetch function. Tries multiple sources in order:
    1. Official Oregon.gov XLSX export
    2. Tableau CSV export
    3. CAMP eLicensing API
    """
    records = fetch_olcc_via_official_xlsx(official_xlsx_url or OFFICIAL_XLSX_URL)
    if records:
        print(f'[OR OLCC] Got {len(records)} records from official XLSX')
        return records

    # Try Tableau next
    records = fetch_olcc_via_tableau()
    if records:
        print(f'[OR OLCC] Got {len(records)} records from Tableau')
        return records

    # Try eLicensing API
    records = fetch_olcc_via_elicensing()
    if records:
        print(f'[OR OLCC] Got {len(records)} records from eLicensing')
        return records

    print('[OR OLCC] All fetch methods exhausted')
    return []


def insert_records(records, dry_run=False):
    """Insert records into registries table using INSERT OR IGNORE for idempotency."""
    if not records:
        print('[OR OLCC] No records to insert.')
        return {'new': 0, 'existing': 0, 'errors': 0}

    conn = get_db_connection()
    cursor = conn.cursor()

    new_count = 0
    existing_count = 0
    error_count = 0

    for rec in records:
        business_name = rec.get('business_name', '').strip()
        license_number = rec.get('license_number', '').strip()
        city = rec.get('city', '').strip()

        if not business_name:
            error_count += 1
            continue

        if not license_number:
            cursor.execute(
                'SELECT 1 FROM registries WHERE business_name = ? AND city = ? AND registry_source = ? LIMIT 1',
                (business_name, city, REGISTRY_SOURCE),
            )
            if cursor.fetchone():
                existing_count += 1
                continue

        if dry_run:
            if license_number:
                cursor.execute(
                    'SELECT 1 FROM registries WHERE license_number = ? AND registry_source = ? LIMIT 1',
                    (license_number, REGISTRY_SOURCE),
                )
                if cursor.fetchone():
                    existing_count += 1
                else:
                    new_count += 1
            else:
                new_count += 1
            continue

        try:
            cursor.execute(
                '''INSERT OR IGNORE INTO registries
                   (business_name, license_number, license_type, license_status,
                    address, city, state, zip, county,
                    registry_source, segment, raw_data, imported_at)
                   VALUES (?, ?, ?, ?, ?, ?, 'OR', ?, ?,
                           ?, ?, ?, CURRENT_TIMESTAMP)''',
                (
                    business_name,
                    license_number or None,
                    rec.get('license_type', ''),
                    rec.get('license_status', 'active'),
                    rec.get('address', ''),
                    city,
                    rec.get('zip', ''),
                    rec.get('county', ''),
                    REGISTRY_SOURCE,
                    SEGMENT,
                    rec.get('raw_data', '{}'),
                ),
            )
            if cursor.rowcount > 0:
                new_count += 1
            else:
                existing_count += 1
        except sqlite3.Error as e:
            print(f'[OR OLCC] DB error for {business_name}: {e}')
            error_count += 1

    if not dry_run:
        conn.commit()
    conn.close()
    return {'new': new_count, 'existing': existing_count, 'errors': error_count}


def _write_summary(summary_path: str, payload: dict):
    if not summary_path:
        return
    out = Path(summary_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2) + '\n', encoding='utf-8')


def import_records(dry_run=False, summary_json='', fail_on_empty_fetch=True, official_xlsx_url=''):
    """Main import entry point."""
    print(f'[OR OLCC] Starting import (dry_run={dry_run})')

    # Ensure DB schema is up to date
    migrate_db()

    records = fetch_records(official_xlsx_url=official_xlsx_url)
    fetched_count = len(records)

    if not records:
        print('[OR OLCC] ⚠️  No records fetched. Possible causes:')
        print('           - Tableau endpoint requires authentication')
        print('           - CAMP eLicensing API changed')
        print('           - Network error')
        print('           Try running with browser scrape or manual CSV download.')
        summary = {
            'importer': 'or_olcc',
            'dry_run': dry_run,
            'fetched': fetched_count,
            'new': 0,
            'existing': 0,
            'errors': 1,
            'blocked': False,
            'status': 'failed_empty_fetch',
        }
        _write_summary(summary_json, summary)
        return 2 if fail_on_empty_fetch else 0

    print(f'[OR OLCC] Fetched {len(records)} cultivator records')

    stats = insert_records(records, dry_run=dry_run)

    print(f'\n[OR OLCC] Import complete (dry_run={dry_run}):')
    print(f'  ✅ New records:     {stats["new"]}')
    print(f'  ⏭️  Already existed: {stats["existing"]}')
    print(f'  ❌ Errors:          {stats["errors"]}')

    if dry_run:
        for rec in records[:3]:
            print(f'  Sample: {rec["business_name"]} | {rec["license_number"]} | {rec["city"]}')

    summary = {
        'importer': 'or_olcc',
        'dry_run': dry_run,
        'fetched': fetched_count,
        'new': stats['new'],
        'existing': stats['existing'],
        'errors': stats['errors'],
        'blocked': False,
        'status': 'ok',
    }
    _write_summary(summary_json, summary)
    return 0


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Import Oregon OLCC cannabis cultivator licenses')
    parser.add_argument('--dry-run', action='store_true', help='Print stats without writing to DB')
    parser.add_argument('--summary-json', '--json', dest='summary_json', default='', help='Write machine-readable run summary JSON')
    parser.add_argument('--official-xlsx-url', default=os.environ.get('OR_OLCC_XLSX_URL', OFFICIAL_XLSX_URL), help='Official Oregon.gov XLSX source URL')
    parser.add_argument('--allow-empty-fetch', action='store_true', help='Do not exit non-zero when fetch returns zero records')
    args = parser.parse_args()
    raise SystemExit(import_records(
        dry_run=args.dry_run,
        summary_json=args.summary_json,
        fail_on_empty_fetch=not args.allow_empty_fetch,
        official_xlsx_url=args.official_xlsx_url,
    ))
